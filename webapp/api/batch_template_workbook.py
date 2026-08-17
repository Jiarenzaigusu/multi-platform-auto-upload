# -*- coding: utf-8 -*-
"""批量 Excel 模板的通用工作簿渲染器。

内容模块只声明自己的字段、示例与下拉选项；本模块负责一致的 Excel 外观和序列化，
避免平台或内容类型之间相互依赖。
"""
from __future__ import annotations

from io import BytesIO
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


BatchColumn = tuple[str, str, bool]
ListValidation = tuple[str, Sequence[str], str, str]
MAX_DATA_ROWS = 200


def _column_width(field: str) -> int:
    return {
        "video_path": 26,
        "cover_image_path": 26,
        "image_paths": 48,
        "image_folder_path": 48,
        "title": 24,
        "description": 36,
        "tags": 22,
        "goods_id": 22,
        "activity_topic": 18,
        "music_name": 18,
        "schedule": 20,
        "creator_declaration": 18,
        "original": 12,
    }.get(field, 16)


def _list_validation(values: Sequence[str], error_title: str, error_message: str) -> DataValidation:
    return DataValidation(
        type="list",
        formula1='"' + ",".join(values) + '"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle=error_title,
        error=error_message,
        showInputMessage=False,
    )


def build_content_template(
    *,
    sheet_title: str,
    columns: Sequence[BatchColumn],
    sample_row: Sequence[str],
    list_validations: Sequence[ListValidation] = (),
) -> bytes:
    """根据单一内容模块的声明生成 Excel 模板字节。"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title
    headers = [label for _field, label, _required in columns]
    worksheet.append(headers)
    worksheet.append(list(sample_row))

    for column_index, (field, _label, _required) in enumerate(columns, start=1):
        letter = get_column_letter(column_index)
        worksheet.column_dimensions[letter].width = _column_width(field)
        worksheet.column_dimensions[letter].auto_size = False
        worksheet.cell(1, column_index).font = Font(bold=True)
        worksheet.cell(1, column_index).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        worksheet.cell(2, column_index).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=False
        )
        if field == "goods_id":
            for row_index in range(2, MAX_DATA_ROWS + 2):
                # 商品 ID 是标识符而不是可计算数值；文本格式可保留前导零，避免
                # WPS/Excel 用科学计数法或井号显示长 ID。
                worksheet.cell(row_index, column=column_index).number_format = "@"
        elif field == "schedule":
            for row_index in range(2, MAX_DATA_ROWS + 2):
                worksheet.cell(row_index, column=column_index).number_format = "@"

    worksheet.row_dimensions[1].height = 22
    worksheet.row_dimensions[2].height = 18
    worksheet.sheet_view.showGridLines = True

    field_positions = {field: index for index, (field, _label, _required) in enumerate(columns, start=1)}
    for field, values, error_title, error_message in list_validations:
        column_index = field_positions.get(field)
        if column_index is None:
            continue
        validation = _list_validation(values, error_title, error_message)
        letter = get_column_letter(column_index)
        validation.add(f"{letter}2:{letter}{MAX_DATA_ROWS + 1}")
        worksheet.add_data_validation(validation)

    output = BytesIO()
    try:
        workbook.save(output)
        return output.getvalue()
    finally:
        workbook.close()
