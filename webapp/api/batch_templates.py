from __future__ import annotations

from io import BytesIO
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from webapp.api.batch_jd import JD_BATCH_COLUMNS, JD_SAMPLE_ROW
from webapp.api.batch_tmall import TMALL_BATCH_COLUMNS, TMALL_SAMPLE_ROW
from webapp.api.models import CREATOR_DECLARATIONS


MAX_DATA_ROWS = 200


def _column_widths(columns: Sequence[tuple[str, str, bool]]) -> dict[int, int]:
    hints = {
        "video_path": 26,
        "title": 24,
        "description": 36,
        "tags": 22,
        "goods_id": 16,
        "activity_topic": 18,
        "music_name": 18,
        "schedule": 20,
        "creator_declaration": 18,
        "original": 12,
    }
    return {
        index: hints.get(field, 16)
        for index, (field, _label, _required) in enumerate(columns, start=1)
    }


def _list_validation(formula: str, title: str, message: str) -> DataValidation:
    return DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle=title,
        error=message,
        showInputMessage=False,
    )


def _build_template(
    *,
    sheet_title: str,
    columns: Sequence[tuple[str, str, bool]],
    sample_row: Sequence[str],
    validations: Iterable[tuple[str, DataValidation]],
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title
    headers = [label for _field, label, _required in columns]
    worksheet.append(headers)
    worksheet.append(list(sample_row))
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    for index, _header in enumerate(headers, start=1):
        worksheet.cell(1, index).font = header_font
        worksheet.cell(1, index).alignment = header_alignment
        worksheet.cell(2, index).alignment = data_alignment
    worksheet.row_dimensions[1].height = 22
    worksheet.row_dimensions[2].height = 18
    worksheet.sheet_view.showGridLines = True
    for index, width in _column_widths(columns).items():
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for index, (field, _label, _required) in enumerate(columns, start=1):
        if field == "schedule":
            for row in range(2, MAX_DATA_ROWS + 2):
                worksheet.cell(row, index).number_format = "@"
    for field_name, validation in validations:
        for index, (field, _label, _required) in enumerate(columns, start=1):
            if field == field_name:
                letter = get_column_letter(index)
                validation.add(f"{letter}2:{letter}{MAX_DATA_ROWS + 1}")
                worksheet.add_data_validation(validation)
                break
    output = BytesIO()
    try:
        workbook.save(output)
    finally:
        workbook.close()
    return output.getvalue()


def build_tmall_template() -> bytes:
    creator = _list_validation(
        '"' + ",".join(CREATOR_DECLARATIONS) + '"',
        "无效的创作者声明",
        "请从下拉列表中选择预定义的创作者声明",
    )
    return _build_template(
        sheet_title="天猫批量发布",
        columns=TMALL_BATCH_COLUMNS,
        sample_row=TMALL_SAMPLE_ROW,
        validations=[("creator_declaration", creator)],
    )


def build_jd_template() -> bytes:
    original = _list_validation('"是,否"', "无效的自主原创", '请填写"是"或"否"')
    creator = _list_validation(
        '"' + ",".join(CREATOR_DECLARATIONS) + '"',
        "无效的创作者声明",
        "请从下拉列表中选择预定义的创作者声明",
    )
    return _build_template(
        sheet_title="京东批量发布",
        columns=JD_BATCH_COLUMNS,
        sample_row=JD_SAMPLE_ROW,
        validations=[("original", original), ("creator_declaration", creator)],
    )


def build_batch_template(platform: str) -> bytes:
    builders = {"tmall": build_tmall_template, "jd": build_jd_template}
    try:
        return builders[platform]()
    except KeyError as exc:
        raise ValueError(f"不支持的平台: {platform}") from exc
