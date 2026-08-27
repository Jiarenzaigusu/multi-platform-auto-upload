# -*- coding: utf-8 -*-
"""webapp.ai_copy.router 模块：AI 文案的 FastAPI 路由。

挂载在 /api/ai-copy 前缀下，提供：
- GET  /options: 获取风格/场景/节日选项与 LLM 状态
- GET  /selling-point-template: 下载商品核心卖点 Excel 模板
- POST /selling-point-catalog: 上传卖点 Excel
- DELETE /selling-point-catalog/{id}: 删除卖点目录
- POST /product-references: 读取商品链接资料（预览）
- POST /generate: 生成文案
- POST /import-to-batch-excel: 将生成的文案标题和商品ID导入批量发布Excel
"""
from __future__ import annotations

import asyncio
from copy import copy
import re
from collections.abc import Callable
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from webapp.ai_copy.contracts import (
    FESTIVAL_SUGGESTIONS,
    GenerateCopyRequest,
    GenerateCopyResponse,
    ProductReference,
    ProductReferencesRequest,
    SCENE_LABELS,
    SellingPointCatalogUploadResponse,
    STYLE_LABELS,
)
from webapp.ai_copy.errors import AiCopyError
from webapp.ai_copy.service import AiCopyService
from webapp.llm_adapter.errors import LLMAdapterError


def _normalize_header(value) -> str:
    """归一化表头：去空白与下划线，转小写。用于表头别名匹配。"""
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"[\s_]+", "", text).lower()


def _cell_text(value) -> str:
    """将 Excel 单元格值转换为字符串。"""
    if value is None:
        return ""
    from datetime import date, datetime
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d 00:00")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def create_ai_copy_router(
    service: AiCopyService | Callable[[Request], AiCopyService],
) -> APIRouter:
    """创建 AI 文案 APIRouter。

    :param service: AI 文案服务实例
    :returns: 挂载在 /api/ai-copy 前缀的 APIRouter
    """
    router = APIRouter(prefix="/api/ai-copy", tags=["ai-copy"])

    def resolve(request: Request) -> AiCopyService:
        return service(request) if callable(service) else service

    @router.get("/options")
    def get_options(request: Request = None) -> dict:
        """返回风格/场景/节日选项与当前 LLM 状态。"""
        return {
            "styles": [
                {"value": value.value, "label": label}
                for value, label in STYLE_LABELS.items()
            ],
            "scenes": [
                {"value": value.value, "label": label}
                for value, label in SCENE_LABELS.items()
            ],
            "festivals": list(FESTIVAL_SUGGESTIONS),
            "llm": {
                "ready": resolve(request).llm_ready,
                "model": resolve(request).model,
                "provider": resolve(request).provider_label,
            },
        }

    @router.post(
        "/selling-point-catalog",
        response_model=SellingPointCatalogUploadResponse,
    )
    async def upload_selling_point_catalog(
        file: UploadFile = File(...),
        request: Request = None,
    ) -> SellingPointCatalogUploadResponse:
        """上传商品核心卖点 Excel，返回目录 ID 与解析出的条目。"""
        try:
            selected = resolve(request)
            content = await file.read(selected.max_selling_point_workbook_bytes + 1)
            return await asyncio.to_thread(
                selected.upload_selling_points,
                file.filename or "",
                content,
            )
        except (AiCopyError, LLMAdapterError) as exc:
            raise HTTPException(status_code=exc.status_code, detail=str(exc)) from exc
        finally:
            await file.close()

    @router.get("/selling-point-template")
    def download_selling_point_template() -> StreamingResponse:
        """下载商品核心卖点 Excel 模板。"""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "商品核心卖点"
        worksheet.append(["商品ID或货号", "商品核心内容卖点"])
        worksheet.append(["SKU-001", "轻量透气，适合日常通勤与周末出行"])
        worksheet.append(["", ""])
        worksheet.append(["填写说明", "商品ID或货号不可重复；每行填写一条核心卖点"])
        worksheet.column_dimensions["A"].width = 24
        worksheet.column_dimensions["B"].width = 60
        for cell in worksheet[1]:
            font = copy(cell.font)
            font.bold = True
            cell.font = font
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            for cell in row:
                alignment = copy(cell.alignment)
                alignment.vertical = "top"
                alignment.wrap_text = True
                cell.alignment = alignment
        worksheet.row_dimensions[1].height = 24
        worksheet.row_dimensions[4].height = 34
        worksheet.freeze_panes = "A2"

        output = BytesIO()
        workbook.save(output)
        workbook.close()
        output.seek(0)
        filename = "商品核心卖点模板.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": (
                    'attachment; filename="selling-point-template.xlsx"; '
                    f"filename*=UTF-8''{quote(filename)}"
                )
            },
        )

    @router.delete("/selling-point-catalog/{catalog_id}")
    def delete_selling_point_catalog(
        catalog_id: str, request: Request = None
    ) -> dict[str, bool]:
        """删除指定卖点目录。"""
        return {"deleted": resolve(request).delete_selling_point_catalog(catalog_id)}

    @router.post("/product-references", response_model=list[ProductReference])
    async def inspect_products(
        payload: ProductReferencesRequest,
        request: Request = None,
    ) -> list[ProductReference]:
        """读取商品链接资料（预览，不生成文案）。"""
        try:
            return await asyncio.to_thread(resolve(request).inspect_products, payload)
        except (AiCopyError, LLMAdapterError) as exc:
            raise HTTPException(status_code=exc.status_code, detail=str(exc)) from exc

    @router.post("/generate", response_model=GenerateCopyResponse)
    async def generate_copy(
        payload: GenerateCopyRequest, request: Request = None
    ) -> GenerateCopyResponse:
        """生成文案：匹配卖点 → 读取商品 → LLM 生成 → 校验。"""
        try:
            return await asyncio.to_thread(resolve(request).generate, payload)
        except (AiCopyError, LLMAdapterError) as exc:
            raise HTTPException(status_code=exc.status_code, detail=str(exc)) from exc

    @router.post("/import-to-batch-excel")
    async def import_to_batch_excel(
        file: UploadFile = File(...),
        title: str = Form(""),
        body: str = Form(""),
        product_identifiers: str = Form(""),
    ) -> StreamingResponse:
        """将生成的文案（标题+正文）导入批量发布 Excel 表格。

        按商品 ID/货号组在表格中唯一匹配行：
        - 单个 ID 仅匹配同样的单 ID 行；多个 ID 仅匹配 ID 集合相同的多 ID 行
        - 标题、文案列按实际表头识别，存在才填写
        - 未匹配时在表尾新建一行，仅填写商品 ID/货号、标题和文案列

        :param file: 批量发布 Excel 文件（.xlsx）
        :param title: AI 生成的文案标题
        :param body: AI 生成的文案正文
        :param product_identifiers: 商品 ID/货号（逗号/换行/分号分隔）
        :returns: 修改后的 Excel 文件流
        """
        original_name = (file.filename or "batch.xlsx").strip()
        if not original_name.lower().endswith(".xlsx"):
            await file.close()
            raise HTTPException(status_code=422, detail="请上传 .xlsx 格式的批量发布表格")

        # 限制文件大小（10MB）
        max_bytes = 10 * 1024 * 1024
        try:
            content = await file.read(max_bytes + 1)
        finally:
            await file.close()

        if len(content) > max_bytes:
            raise HTTPException(status_code=413, detail="Excel 文件不能超过 10 MB")

        # 解析商品 ID/货号列表
        identifiers = [
            item.strip()
            for item in re.split(r"[\s,，;；\n]+", product_identifiers)
            if item.strip()
        ]
        if not identifiers:
            raise HTTPException(status_code=422, detail="未提供商品 ID/货号，无法定位导入行")

        if not title and not body:
            raise HTTPException(status_code=422, detail="标题与文案均为空，没有可导入的内容")

        try:
            # 在线程中执行 Excel 操作（避免阻塞事件循环）
            modified_content, stats = await asyncio.to_thread(
                _import_copy_to_excel,
                content,
                title,
                body,
                identifiers,
            )
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"处理 Excel 文件失败：{exc}") from exc

        # 汇总信息：仅用 ASCII 写入响应头（HTTP 头禁止非 ASCII 字符），
        # 前端解析后再拼成中文提示。
        summary = f"updated={stats['matched']};created={stats['created']}"

        return StreamingResponse(
            iter([modified_content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "X-Import-Summary": summary,
                "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
                "Pragma": "no-cache",
                "Expires": "0",
            },
        )

    return router


def _split_goods_identifiers(value: str) -> list[str]:
    """拆分商品单元格中的多个 ID，兼容逗号、分号、空格与单元格内换行。"""
    return [
        item.strip()
        for item in re.split(r"[\s,，;；]+", value)
        if item.strip()
    ]


def _identifier_group_key(identifiers: list[str]) -> tuple[str, ...]:
    """构造无序且去重的 ID 组键，供整组唯一匹配使用。"""
    normalized: set[str] = set()
    for identifier in identifiers:
        value = identifier.strip().lower()
        if not value:
            continue
        # Excel 把纯数字 ID 当作数字保存时，去除前导零后仍可匹配。
        normalized.add(str(int(value)) if value.isdigit() else value)
    return tuple(sorted(normalized))


# 表头归一化后的别名集合（用于灵活识别商品 ID/货号、标题、文案列）
_GOODS_ID_ALIASES = {
    "商品id", "商品编号", "goodsid", "货号", "货品id",
    "skuid", "sku", "产品id", "产品编号", "商品货号",
}
_TITLE_ALIASES = {
    "标题", "title", "商品标题", "宝贝标题",
}
_BODY_ALIASES = {
    "文案", "发布文案", "description", "正文", "内容", "详情",
    "商品文案", "正文文案", "body",
}


def _locate_columns(rows: list, header_scan_limit: int = 10):
    """在前若干行中查找表头行，返回目标列的 1-based 索引。

    逐行扫描前 header_scan_limit 行；识别到「商品ID/货号」列即视为表头行。
    「标题」和「文案」列均为可选，按实际表头决定是否写入。

    :returns: (header_row_index, goods_col, title_col, body_col)
    """
    for row_idx, row in enumerate(rows[:header_scan_limit], start=1):
        normalized = [_normalize_header(cell) for cell in row]
        goods_col = title_col = body_col = None
        for col_idx, header in enumerate(normalized):
            if header in _GOODS_ID_ALIASES:
                goods_col = col_idx + 1
            elif header in _TITLE_ALIASES:
                title_col = col_idx + 1
            elif header in _BODY_ALIASES:
                body_col = col_idx + 1
        if goods_col is not None:
            return row_idx, goods_col, title_col, body_col
    return None, None, None, None


def _last_populated_row(worksheet, header_row_index: int) -> int:
    """返回实际包含数据的最后一行，忽略模板预设的格式和下拉验证范围。"""
    for row_idx in range(worksheet.max_row, header_row_index, -1):
        if any(
            _cell_text(cell.value)
            for cell in worksheet[row_idx]
        ):
            return row_idx
    return header_row_index


def _import_copy_to_excel(
    content: bytes,
    title: str,
    body: str,
    identifiers: list[str],
) -> bytes:
    """将文案按商品 ID/货号组导入 Excel。

    - 标题列、文案列仅当表头中存在时才写入；不存在则跳过对应字段
      （如京东模板只有「标题」列则只填标题，天猫模板有「标题」「文案」则都填）。
    - 商品 ID/货号组作为唯一键：单 ID 只匹配单 ID 行；多 ID 只匹配相同 ID 集合的行。
      ID 的顺序、逗号/换行分隔形式不影响匹配。
    - 未匹配时新增一行，只写入商品 ID/货号与实际存在的标题、文案列。

    :param content: 原始 Excel 字节
    :param title: 标题文本
    :param body: 文案正文
    :param identifiers: 商品 ID/货号列表
    :returns: (修改后的 Excel 字节, 统计信息 {"matched": 更新行数, "created": 新建行数})
    """
    from zipfile import BadZipFile

    try:
        workbook = load_workbook(BytesIO(content))
    except (BadZipFile, Exception) as exc:
        raise ValueError("无法读取 Excel 文件，请确保是有效的 .xlsx 格式") from exc

    try:
        worksheet = workbook.active
        if worksheet is None:
            raise ValueError("Excel 文件中没有工作表")

        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel 文件为空，无法导入")

        header_row_index, goods_col, title_col, body_col = _locate_columns(rows)
        if header_row_index is None:
            raise ValueError("未找到商品 ID/货号表头，无法确定导入行")
        if goods_col is None:
            raise ValueError("未找到「商品ID/货号」列，无法确定导入到哪一行")
        if title_col is None and body_col is None:
            raise ValueError("未找到可导入的「标题」或「文案」表头")

        data_start = header_row_index + 1
        last_populated_row = _last_populated_row(worksheet, header_row_index)

        target_group = _identifier_group_key(identifiers)
        if not target_group:
            raise ValueError("未提供有效的商品 ID/货号，无法定位导入行")

        matched_row = None
        for row_idx in range(data_start, last_populated_row + 1):
            row_identifiers = _split_goods_identifiers(
                _cell_text(worksheet.cell(row=row_idx, column=goods_col).value)
            )
            if _identifier_group_key(row_identifiers) == target_group:
                matched_row = row_idx
                break

        created_count = 0
        if matched_row is None:
            # 不使用 worksheet.max_row：模板样式/下拉验证会把它扩展到数百行。
            matched_row = last_populated_row + 1
            worksheet.cell(
                row=matched_row,
                column=goods_col,
                value="\n".join(dict.fromkeys(item.strip() for item in identifiers if item.strip())),
            )
            created_count = 1

        if title_col is not None:
            worksheet.cell(row=matched_row, column=title_col, value=title)
        if body_col is not None:
            worksheet.cell(row=matched_row, column=body_col, value=body)

        output = BytesIO()
        workbook.save(output)
        stats = {"matched": 0 if created_count else 1, "created": created_count}
        return output.getvalue(), stats
    finally:
        workbook.close()
