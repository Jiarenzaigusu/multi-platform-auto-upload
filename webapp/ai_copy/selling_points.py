# -*- coding: utf-8 -*-
"""webapp.ai_copy.selling_points 模块：商品核心卖点 Excel 解析与内存目录存储。

卖点表格仅保留在当前服务内存（不写入运行目录），通过 SellingPointCatalogStore
管理，支持 TTL 过期与 LRU 淘汰。前端上传后获得 catalog_id，后续生成文案时引用。
"""
from __future__ import annotations

import re
import secrets
import threading
import time
from collections import OrderedDict
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from webapp.ai_copy.contracts import (
    SellingPointCatalogUploadResponse,
    SellingPointReference,
)
from webapp.ai_copy.errors import (
    SellingPointCatalogError,
    SellingPointCatalogNotFoundError,
)


# "商品ID或货号"列的可识别表头（归一化后匹配）
IDENTIFIER_HEADERS = {
    "商品id或货号",
    "商品id或者货号",
    "商品id货号",
    "商品id",
    "商品编号",
    "货号",
    "productid",
    "sku",
}
# "商品核心内容卖点"列的可识别表头
SELLING_POINT_HEADERS = {
    "商品核心内容卖点",
    "商品的核心内容卖点",
    "核心内容卖点",
    "商品核心卖点",
    "核心卖点",
    "商品卖点",
    "sellingpoint",
}


def _cell_text(value: Any) -> str:
    """将 Excel 单元格值转为字符串（布尔值转空）。"""
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalized_header(value: Any) -> str:
    """归一化表头：去空白与标点，转小写，用于表头匹配。"""
    return re.sub(r"[\s_/（）()-]+", "", _cell_text(value)).casefold()


def _compact_selling_point(value: Any) -> str:
    """压缩卖点文本中的连续空白为单个空格。"""
    return re.sub(r"\s+", " ", _cell_text(value)).strip()


def _catalog_key(identifier: str) -> str:
    """生成目录查找键（去空白 + 小写，用于精确匹配）。"""
    return identifier.strip().casefold()


def parse_selling_point_workbook(
    content: bytes,
    *,
    max_rows: int,
) -> tuple[SellingPointReference, ...]:
    """解析卖点 Excel，返回条目元组。

    流程：
    1. 打开工作簿
    2. 在前 10 行查找两个必填表头（"商品ID或货号" + "商品核心内容卖点"）
    3. 逐行解析，校验非空、长度、ID 唯一
    4. 超过 max_rows 报错

    :returns: SellingPointReference 元组
    :raises SellingPointCatalogError: 解析失败
    """
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, EOFError, InvalidFileException, KeyError, OSError, ValueError) as exc:
        raise SellingPointCatalogError("无法读取 Excel，请上传有效的 .xlsx 文件") from exc

    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header_positions: tuple[int, int] | None = None
        header_row_number = 0

        # 在前 10 行查找表头
        for row_number, values in enumerate(rows, start=1):
            identifier_index = None
            selling_point_index = None
            for index, value in enumerate(values):
                normalized = _normalized_header(value)
                if normalized in IDENTIFIER_HEADERS:
                    identifier_index = index
                if normalized in SELLING_POINT_HEADERS:
                    selling_point_index = index
            if identifier_index is not None and selling_point_index is not None:
                if identifier_index == selling_point_index:
                    raise SellingPointCatalogError("Excel 两个必填表头不能使用同一列")
                header_positions = (identifier_index, selling_point_index)
                header_row_number = row_number
                break
            if row_number >= 10:
                break

        if header_positions is None:
            raise SellingPointCatalogError(
                "Excel 必须包含“商品ID或货号”和“商品核心内容卖点”两列"
            )

        identifier_index, selling_point_index = header_positions
        entries: list[SellingPointReference] = []
        seen: dict[str, int] = {}
        # 逐行解析数据
        for row_number, values in enumerate(rows, start=header_row_number + 1):
            identifier = (
                _cell_text(values[identifier_index])
                if identifier_index < len(values)
                else ""
            )
            selling_point = (
                _compact_selling_point(values[selling_point_index])
                if selling_point_index < len(values)
                else ""
            )
            if not identifier and not selling_point:
                continue
            if not identifier:
                raise SellingPointCatalogError(f"Excel 第 {row_number} 行缺少商品 ID 或货号")
            if not selling_point:
                raise SellingPointCatalogError(f"Excel 第 {row_number} 行缺少商品核心内容卖点")
            if len(identifier) > 100:
                raise SellingPointCatalogError(f"Excel 第 {row_number} 行商品 ID 或货号超过 100 个字符")
            if len(selling_point) > 2000:
                raise SellingPointCatalogError(f"Excel 第 {row_number} 行核心卖点超过 2000 个字符")

            # 校验 ID 唯一性
            key = _catalog_key(identifier)
            if key in seen:
                raise SellingPointCatalogError(
                    f"商品 ID 或货号“{identifier}”在 Excel 第 {seen[key]} 行和第 {row_number} 行重复"
                )
            seen[key] = row_number
            entries.append(
                SellingPointReference(identifier=identifier, selling_point=selling_point)
            )
            if len(entries) > max_rows:
                raise SellingPointCatalogError(f"单个 Excel 最多支持 {max_rows} 条商品卖点")

        if not entries:
            raise SellingPointCatalogError("Excel 中至少需要一条商品卖点数据")
        return tuple(entries)
    finally:
        workbook.close()


@dataclass(frozen=True, slots=True)
class _SellingPointCatalog:
    """内存中的卖点目录条目（不可变）。"""

    catalog_id: str
    filename: str
    entries: tuple[SellingPointReference, ...]
    by_identifier: dict[str, SellingPointReference]  # identifier_key → entry
    expires_at: float


class SellingPointCatalogStore:
    """卖点目录的内存存储，支持 TTL 过期与 LRU 淘汰。

    所有目录仅存在内存中，不写入运行目录；服务重启后需重新上传。
    """

    def __init__(
        self,
        *,
        max_workbook_bytes: int = 5 * 1024 * 1024,
        max_rows: int = 5000,
        ttl_seconds: float = 12 * 60 * 60,
        max_catalogs: int = 8,
    ) -> None:
        """初始化卖点目录存储。

        :param max_workbook_bytes: 单个 Excel 最大字节数
        :param max_rows: 单个 Excel 最多行数
        :param ttl_seconds: 目录 TTL（默认 12 小时）
        :param max_catalogs: 内存最多保留多少目录
        """
        self.max_workbook_bytes = max_workbook_bytes
        self.max_rows = max_rows
        self.ttl_seconds = ttl_seconds
        self.max_catalogs = max_catalogs
        self._lock = threading.RLock()
        self._catalogs: OrderedDict[str, _SellingPointCatalog] = OrderedDict()

    def upload(self, filename: str, content: bytes) -> SellingPointCatalogUploadResponse:
        """上传卖点 Excel，解析后存入内存目录。

        :returns: 上传响应（含 catalog_id 与全部条目）
        """
        # 安全化文件名
        safe_filename = Path((filename or "").replace("\\", "/")).name[:255]
        if not safe_filename.lower().endswith(".xlsx"):
            raise SellingPointCatalogError("卖点表格仅支持 .xlsx 格式")
        if not content:
            raise SellingPointCatalogError("上传的 Excel 文件为空")
        if len(content) > self.max_workbook_bytes:
            max_mib = self.max_workbook_bytes // (1024 * 1024)
            raise SellingPointCatalogError(f"卖点 Excel 不能超过 {max_mib} MiB")

        entries = parse_selling_point_workbook(content, max_rows=self.max_rows)
        catalog_id = secrets.token_urlsafe(24)
        catalog = _SellingPointCatalog(
            catalog_id=catalog_id,
            filename=safe_filename,
            entries=entries,
            by_identifier={_catalog_key(item.identifier): item for item in entries},
            expires_at=time.monotonic() + self.ttl_seconds,
        )
        with self._lock:
            self._prune_locked()
            # 达到上限时淘汰最旧的（LRU）
            while len(self._catalogs) >= self.max_catalogs:
                self._catalogs.popitem(last=False)
            self._catalogs[catalog_id] = catalog

        return SellingPointCatalogUploadResponse(
            catalog_id=catalog_id,
            filename=safe_filename,
            row_count=len(entries),
            entries=list(entries),
        )

    def resolve(
        self,
        catalog_id: str,
        identifiers: list[str],
    ) -> list[SellingPointReference]:
        """按 catalog_id 与商品 ID 列表解析卖点条目。

        :raises SellingPointCatalogNotFoundError: 目录不存在或已过期
        :raises SellingPointCatalogError: 有商品 ID 在目录中找不到
        """
        with self._lock:
            self._prune_locked()
            catalog = self._catalogs.get(catalog_id)
            if catalog is None:
                raise SellingPointCatalogNotFoundError(
                    "卖点 Excel 已失效或服务已重启，请重新上传"
                )
            # LRU：移动到末尾
            self._catalogs.move_to_end(catalog_id)

            # 检查缺失的商品 ID
            missing = [
                identifier
                for identifier in identifiers
                if _catalog_key(identifier) not in catalog.by_identifier
            ]
            if missing:
                raise SellingPointCatalogError(
                    "Excel 中未找到以下商品 ID 或货号：" + "、".join(missing)
                )
            return [
                catalog.by_identifier[_catalog_key(identifier)]
                for identifier in identifiers
            ]

    def delete(self, catalog_id: str) -> bool:
        """删除指定目录，返回是否删除成功。"""
        with self._lock:
            return self._catalogs.pop(catalog_id, None) is not None

    def _prune_locked(self) -> None:
        """在持锁状态下清理过期目录。"""
        now = time.monotonic()
        expired = [
            catalog_id
            for catalog_id, catalog in self._catalogs.items()
            if catalog.expires_at <= now
        ]
        for catalog_id in expired:
            self._catalogs.pop(catalog_id, None)
